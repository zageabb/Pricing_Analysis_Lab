from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(slots=True)
class SpreadsheetData:
    file_name: str
    source_path: Path | None
    sheet_name: str | None
    header_row: int
    sheet_names: list[str]
    columns: list[str]
    rows: list[dict[str, Any]]

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def column_count(self) -> int:
        return len(self.columns)

    def preview(self, limit: int = 10) -> list[dict[str, Any]]:
        return self.rows[:limit]


def load_spreadsheet(
    path: str | Path,
    sheet_name: str | None = None,
    header_row: int = 1,
) -> SpreadsheetData:
    source_path = Path(path)
    suffix = source_path.suffix.lower()
    if header_row < 1:
        raise ValueError("Header row must be 1 or greater.")
    if suffix == ".csv":
        return _load_csv(source_path, header_row=header_row)
    if suffix == ".xlsx":
        return _load_xlsx(source_path, sheet_name=sheet_name, header_row=header_row)
    raise ValueError(f"Unsupported file type: {suffix}")


def _load_csv(path: Path, header_row: int) -> SpreadsheetData:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)

    columns, records = _extract_records(rows, header_row=header_row)
    return SpreadsheetData(
        file_name=path.name,
        source_path=path,
        sheet_name=None,
        header_row=header_row,
        sheet_names=[],
        columns=columns,
        rows=records,
    )


def _load_xlsx(path: Path, sheet_name: str | None = None, header_row: int = 1) -> SpreadsheetData:
    workbook = load_workbook(path, data_only=True)
    sheet_names = workbook.sheetnames
    active_sheet_name = sheet_name or sheet_names[0]

    if active_sheet_name not in sheet_names:
        raise ValueError(f"Sheet '{active_sheet_name}' was not found.")

    worksheet = workbook[active_sheet_name]
    raw_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    columns, records = _extract_records(raw_rows, header_row=header_row)
    return SpreadsheetData(
        file_name=path.name,
        source_path=path,
        sheet_name=active_sheet_name,
        header_row=header_row,
        sheet_names=sheet_names,
        columns=columns,
        rows=records,
    )


def _extract_records(raw_rows: list[list[Any]], header_row: int) -> tuple[list[str], list[dict[str, Any]]]:
    if not raw_rows:
        raise ValueError("Spreadsheet is empty.")
    header_index = header_row - 1
    if header_index >= len(raw_rows):
        raise ValueError(f"Header row {header_row} is beyond the end of the spreadsheet.")
    columns = _normalize_columns(raw_rows[header_index])
    records = _rows_to_records(columns, raw_rows[header_index + 1 :])
    return columns, records


def _normalize_columns(header_row: list[Any]) -> list[str]:
    columns: list[str] = []
    seen: dict[str, int] = {}
    for index, raw_value in enumerate(header_row, start=1):
        base_name = str(raw_value).strip() if raw_value not in (None, "") else f"column_{index}"
        count = seen.get(base_name, 0)
        seen[base_name] = count + 1
        columns.append(base_name if count == 0 else f"{base_name}_{count + 1}")
    return columns


def _rows_to_records(columns: list[str], raw_rows: list[list[Any]]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for raw_row in raw_rows:
        padded = list(raw_row[: len(columns)]) + [None] * max(0, len(columns) - len(raw_row))
        record = {column: _normalize_cell(value) for column, value in zip(columns, padded, strict=True)}
        if all(value in (None, "") for value in record.values()):
            continue
        records.append(record)
    return records


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        lowered = stripped.lower()
        if lowered in {"true", "false"}:
            return lowered == "true"
        if _looks_like_int(stripped):
            return int(stripped)
        if _looks_like_float(stripped):
            return float(stripped)
        if _looks_like_iso_date(stripped):
            return stripped
        return stripped
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _looks_like_int(value: str) -> bool:
    return value.lstrip("+-").isdigit()


def _looks_like_float(value: str) -> bool:
    if value.count(".") != 1:
        return False
    left, right = value.split(".", 1)
    left = left.lstrip("+-")
    return bool(left or right) and left.isdigit() and right.isdigit()


def _looks_like_iso_date(value: str) -> bool:
    parts = value.split("T", 1)[0].split("-")
    return len(parts) == 3 and all(part.isdigit() for part in parts)
