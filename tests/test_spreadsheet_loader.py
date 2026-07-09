from pathlib import Path

from openpyxl import Workbook

from pricing_analysis_lab.services.dataset_profiler import profile_dataset
from pricing_analysis_lab.services.spreadsheet_loader import load_spreadsheet


def test_load_csv_and_profile_columns(tmp_path: Path):
    csv_path = tmp_path / "pricing.csv"
    csv_path.write_text(
        "supplier,category,price,region\n"
        "Acme,Transformer,100,UK\n"
        "Bravo,Transformer,120,UK\n"
        ",Cable,,DE\n",
        encoding="utf-8",
    )

    dataset = load_spreadsheet(csv_path)
    profile = profile_dataset(dataset)

    assert dataset.columns == ["supplier", "category", "price", "region"]
    assert dataset.row_count == 3
    assert profile["row_count"] == 3
    assert profile["columns"][0]["null_count"] == 1
    assert profile["columns"][2]["inferred_type"] == "numeric"


def test_load_xlsx_with_sheet_selection(tmp_path: Path):
    xlsx_path = tmp_path / "pricing.xlsx"
    workbook = Workbook()
    default_sheet = workbook.active
    default_sheet.title = "Sheet1"
    default_sheet.append(["supplier", "price"])
    default_sheet.append(["Acme", 100])

    second_sheet = workbook.create_sheet("Quotes")
    second_sheet.append(["supplier", "lead_time"])
    second_sheet.append(["Bravo", 14])
    workbook.save(xlsx_path)

    dataset = load_spreadsheet(xlsx_path, sheet_name="Quotes")
    profile = profile_dataset(dataset)

    assert dataset.sheet_name == "Quotes"
    assert dataset.sheet_names == ["Sheet1", "Quotes"]
    assert dataset.columns == ["supplier", "lead_time"]
    assert profile["columns"][1]["inferred_type"] == "numeric"


def test_invalid_sheet_name_raises(tmp_path: Path):
    xlsx_path = tmp_path / "pricing.xlsx"
    workbook = Workbook()
    workbook.active.append(["a"])
    workbook.save(xlsx_path)

    try:
        load_spreadsheet(xlsx_path, sheet_name="Missing")
    except ValueError as exc:
        assert "Missing" in str(exc)
    else:
        raise AssertionError("Expected invalid sheet name to raise ValueError")
